import streamlit as st
import pandas as pd
from datetime import timedelta
from io import BytesIO
import requests
from openpyxl import load_workbook

st.set_page_config(page_title="Отчет ТТК", layout="wide")
st.title("📞 Отчет ТТК — Мэтчинг звонков и визитов (60 минут)")

st.markdown("""
**Инструкция:**
1. Загрузите файл **выгрузки из Метрики** — [открыть](https://metrika.yandex.ru/stat/6cfa6793-da4e-405e-8815-75076218c2af?id=51634436&period=2025-06-01%3A2025-06-11&group=day&currency=RUB&dimension_mode=tree&table=visits)  
   *(Период — с начала месяца до последнего воскресенья, формат — XLSX. Цель — **Mouseover Телефон 2**)  
   [🎓 Тестовый файл](https://github.com/Exepicode/ttk/raw/refs/heads/main/тестовый-метрика)
2. Загрузите файл **звонков из CRM** *(присылается КС в чат под названием «Детальный отчет»)*  
   [🎓 Тестовый файл](https://github.com/Exepicode/ttk/raw/refs/heads/main/тестовый-Детальный%20отчет%20Июнь%201-8.xlsx)
3. Загрузите файл **статистики из Мастера отчетов** — [открыть](https://direct.yandex.ru/dna/reports/wizard?ulogin=ttk-igc&state=157105)  
   [🎓 Тестовый файл](https://github.com/Exepicode/ttk/raw/refs/heads/main/тестовый%20файл%20мастер%20отчетов.xlsx)
4. Заполните показатели **Факт** в блоках Поиск и РСЯ
5. *(Опционально)* Измените плановые значения (по умолчанию — июнь 2025)
6. Нажмите кнопку **«Сгенерировать отчет»**
7. Скачайте готовый отчет
""")

metrika_file = st.file_uploader("📊 Загрузите выгрузку из Метрики", type="xlsx")
calls_file = st.file_uploader("📞 Загрузите звонки из CRM", type="xlsx")
direct_file = st.file_uploader("📈 Загрузите файл из Мастера отчетов (Яндекс.Директ) — [открыть отчет](https://direct.yandex.ru/dna/reports/wizard?ulogin=ttk-igc&state=157105)", type="xlsx")

st.header("🧾 Факт: Ввод данных")

report_date_range = st.date_input("📅 Период отчета", value=(pd.to_datetime("today").replace(day=1), pd.to_datetime("today")), format="DD.MM.YYYY")

st.markdown("### 🔍 Поиск")
col1, col2, col3, col4 = st.columns([1.5, 1, 1, 1])
with col1:
    search_cost = st.number_input("💰 Расход (с НДС)", min_value=0.0, step=100.0)
with col2:
    search_impressions = st.number_input("👁 Показы", min_value=0, step=100)
with col3:
    search_clicks = st.number_input("🖱 Клики", min_value=0, step=1)
with col4:
    search_conversions = st.number_input("📩 Заявки (по Метрике)", min_value=0, step=1)

st.markdown("### 🟡 РСЯ")
col5, col6, col7, col8 = st.columns([1.5, 1, 1, 1])
with col5:
    rsya_cost = st.number_input("💰 Расход (с НДС)", min_value=0.0, step=100.0, key="rsya_cost")
with col6:
    rsya_impressions = st.number_input("👁 Показы", min_value=0, step=100, key="rsya_impressions")
with col7:
    rsya_clicks = st.number_input("🖱 Клики", min_value=0, step=1, key="rsya_clicks")
with col8:
    rsya_conversions = st.number_input("📩 Заявки (по Метрике)", min_value=0, step=1, key="rsya_conversions")

st.markdown("### 🧠 Плановые показатели (необязательно)")
show_plan_inputs = st.checkbox("📋 Изменить плановые значения", value=False)

if show_plan_inputs:
    with st.container():
        st.markdown("⚙️ По умолчанию установлены плановые значения июня 2025. Вы можете изменить их ниже.")
        col_plan1, col_plan2, col_plan3, col_plan4 = st.columns([1.5, 1, 1, 1])
        with col_plan1:
            search_cost_plan = st.number_input("💰 План расход Поиск", min_value=0.0, step=100.0, value=630540.0)
        with col_plan2:
            search_impressions_plan = st.number_input("👁 План показы Поиск", min_value=0, step=100, value=50768)
        with col_plan3:
            search_clicks_plan = st.number_input("🖱 План клики Поиск", min_value=0, step=1, value=7006)
        with col_plan4:
            search_conversions_plan = st.number_input("📩 План заявки Поиск", min_value=0, step=1, value=220)

        col_plan5, col_plan6, col_plan7, col_plan8 = st.columns([1.5, 1, 1, 1])
        with col_plan5:
            rsya_cost_plan = st.number_input("💰 План расход РСЯ", min_value=0.0, step=100.0, value=61008.0)
        with col_plan6:
            rsya_impressions_plan = st.number_input("👁 План показы РСЯ", min_value=0, step=100, value=211833)
        with col_plan7:
            rsya_clicks_plan = st.number_input("🖱 План клики РСЯ", min_value=0, step=1, value=2542)
        with col_plan8:
            rsya_conversions_plan = st.number_input("📩 План заявки РСЯ", min_value=0, step=1, value=22)
else:
    search_cost_plan = 630540.0
    search_impressions_plan = 50768
    search_clicks_plan = 7006
    search_conversions_plan = 220
    rsya_cost_plan = 61008.0
    rsya_impressions_plan = 211833
    rsya_clicks_plan = 2542
    rsya_conversions_plan = 22

def normalize_region(s):
    return str(s).strip().lower().replace('г.', '').replace('-', '').replace('ё', 'е').replace(' ', '')

def process_visits(df):
    for i, row in df.iterrows():
        if str(row.iloc[0]).strip().lower().startswith('дата и время визита'):
            df.columns = row
            df = df.iloc[i+1:]
            break
    df = df.dropna(how='all')
    df = df[~df.iloc[:, 0].astype(str).str.contains('итого', case=False, na=False)]
    df.columns = df.columns.str.strip()
    df['visit_time'] = pd.to_datetime(df['Дата и время визита'], errors='coerce')
    df['region'] = df['Город'].apply(normalize_region)
    df = df.dropna(subset=['visit_time', 'region'])
    df['visit_end'] = df['visit_time'] + timedelta(minutes=60)
    return df

def process_calls(df):
    df.columns = df.columns.str.strip()
    df['call_time'] = pd.to_datetime(df['Дата'].astype(str) + ' ' + df['Время'].astype(str), errors='coerce')
    df['region'] = df['Город'].apply(normalize_region)
    if '№ тел.' in df.columns:
        df = df.rename(columns={'№ тел.': 'Телефон'})
    df = df.dropna(subset=['call_time', 'region'])
    return df

def match_data(calls, visits):
    merged = pd.merge(
        calls[['call_time', 'region']],
        visits[['visit_time', 'visit_end', 'region']],
        on='region',
        how='inner'
    )
    merged = merged[
        (merged['call_time'] >= merged['visit_time']) &
        (merged['call_time'] <= merged['visit_end'])
    ].copy()

    merged['time_diff'] = (merged['call_time'] - merged['visit_time']).abs()
    merged = merged.sort_values('time_diff').drop_duplicates(subset=['visit_time', 'region'])
    merged = merged.drop(columns=['time_diff'])
    merged['Call DateTime'] = merged['call_time']
    return merged[['call_time', 'visit_time', 'region']].rename(columns={
        'call_time': 'Время звонка',
        'visit_time': 'Время визита'
    }).drop_duplicates()

result_df = pd.DataFrame()

if not result_df.empty:
    st.success(f"✅ Найдено совпадений: {len(result_df)}")

if not metrika_file or not calls_file:
    st.info("ℹ️ Отчет будет без совпадений — загрузите оба файла для мэтчинга.")
else:
    st.info("ℹ️ Отчет ожидает генерации — нажмите кнопку ниже.")

if st.button("🚀 Сгенерировать отчет"):
    with st.spinner("🔄 Обрабатываем данные..."):
        try:
            if metrika_file and calls_file:
                visits_raw = pd.read_excel(metrika_file, header=None)
                visits_df = process_visits(visits_raw)
                calls_df = pd.read_excel(calls_file)
                calls_df = process_calls(calls_df)
                result_df = match_data(calls_df, visits_df)
            else:
                result_df = pd.DataFrame()  # пустая таблица совпадений

            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:

                # Добавляем шаблон "Факт" из GitHub как лист
                try:
                    template_url = "https://github.com/Exepicode/ttk/raw/main/ТТК-шаблон-отчета.xlsx"
                    response = requests.get(template_url)
                    if response.status_code == 200:
                        template_excel = BytesIO(response.content)
                        wb_template = load_workbook(template_excel, data_only=False)

                        if wb_template.sheetnames:
                            source_ws = wb_template.worksheets[0]
                            source_ws.title = "Факт"
                            target_ws = writer.book.create_sheet("Факт")

                            for row in source_ws.iter_rows():
                                if all(cell.value in (None, "") for cell in row):
                                    continue
                                for cell in row:
                                    new_cell = target_ws.cell(row=cell.row, column=cell.column)
                                    if cell.data_type == 'f' or (isinstance(cell.value, str) and cell.value.startswith("=")):
                                        new_cell.value = cell.value
                                    else:
                                        new_cell.value = cell.value
                                    try:
                                        if cell.has_style:
                                            new_cell.font = cell.font.copy()
                                            new_cell.border = cell.border.copy()
                                            new_cell.fill = cell.fill.copy()
                                            new_cell.number_format = cell.number_format
                                            new_cell.protection = cell.protection.copy()
                                            new_cell.alignment = cell.alignment.copy()
                                    except Exception as style_error:
                                        st.warning(f"⚠️ Стиль не скопирован для ячейки {cell.coordinate}: {style_error}")

                            for col_letter, dim in source_ws.column_dimensions.items():
                                target_ws.column_dimensions[col_letter].width = dim.width
                            for row_idx, dim in source_ws.row_dimensions.items():
                                target_ws.row_dimensions[row_idx].height = dim.height
                            target_ws.row_dimensions[7].height = 1
                            target_ws.column_dimensions['F'].width = 15
                        else:
                            st.warning("⚠️ В шаблоне отсутствуют листы")
                    else:
                        st.warning(f"⚠️ Не удалось скачать шаблон: статус {response.status_code}")
                except Exception as e:
                    st.warning(f"⚠️ Ошибка при вставке шаблона 'Факт': {e}")

                try:
                    fact_ws = writer.book["Факт"]
                    fact_ws["D4"] = f"{report_date_range[0].strftime('%d.%m.%Y')} – {report_date_range[1].strftime('%d.%m.%Y')}"
                    fact_ws["F8"] = search_cost
                    fact_ws["F9"] = rsya_cost
                    fact_ws["H8"] = search_impressions
                    fact_ws["H9"] = rsya_impressions
                    fact_ws["J8"] = search_clicks
                    fact_ws["J9"] = rsya_clicks
                    fact_ws["P8"] = search_conversions
                    fact_ws["P9"] = rsya_conversions
                    fact_ws["Q8"] = len(result_df)
                    fact_ws["E8"] = search_cost_plan
                    fact_ws["E9"] = rsya_cost_plan
                    fact_ws["G8"] = search_impressions_plan
                    fact_ws["G9"] = rsya_impressions_plan
                    fact_ws["I8"] = search_clicks_plan
                    fact_ws["I9"] = rsya_clicks_plan
                    fact_ws["O8"] = search_conversions_plan
                    fact_ws["O9"] = rsya_conversions_plan
                except Exception as e:
                    st.warning(f"⚠️ Не удалось записать данные в 'Факт': {e}")

                result_df.to_excel(writer, sheet_name="Совпадения", index=False)
                # Устанавливаем ширину столбцов на листе "Совпадения"
                worksheet = writer.sheets["Совпадения"]
                for column_cells in worksheet.columns:
                    max_length = 20
                    col_letter = column_cells[0].column_letter
                    worksheet.column_dimensions[col_letter].width = max_length
                if metrika_file:
                    visits_raw.to_excel(writer, sheet_name="Метрика", index=False, header=False)
                if calls_file:
                    pd.read_excel(calls_file).to_excel(writer, sheet_name="Звонки", index=False)
                if direct_file:
                    pd.read_excel(direct_file).to_excel(writer, sheet_name="Статистика Директ", index=False)

            st.download_button(
                label="📥 Скачать отчет (XLSX)",
                data=output.getvalue(),
                file_name="Отчет_ТТК.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            if search_conversions_plan > 0:
                lead_price_fact = search_cost / search_conversions if search_conversions > 0 else 0
                lead_price_plan = search_cost_plan / search_conversions_plan
                lead_price_diff_pct = ((lead_price_fact - lead_price_plan) / lead_price_plan) * 100 if lead_price_plan else 0

                st.markdown("### 📊 Выводы по показателям")
                st.markdown(f"""
                **Поиск:**
                - Выполнение плана по расходу: {search_cost / search_cost_plan * 100:.1f}%
                - Выполнение плана по показам: {search_impressions / search_impressions_plan * 100:.1f}%
                - Выполнение плана по кликам: {search_clicks / search_clicks_plan * 100:.1f}%
                - Выполнение плана по заявкам: {search_conversions / search_conversions_plan * 100:.1f}%
                - Цена лида (факт): {lead_price_fact:.0f} ₽
                - Цена лида (план): {lead_price_plan:.0f} ₽
                - Отклонение цены лида от плана: {lead_price_diff_pct:+.1f}%
                """)

            if rsya_conversions_plan > 0:
                lead_price_fact_rsya = rsya_cost / rsya_conversions if rsya_conversions > 0 else 0
                lead_price_plan_rsya = rsya_cost_plan / rsya_conversions_plan
                lead_price_diff_pct_rsya = ((lead_price_fact_rsya - lead_price_plan_rsya) / lead_price_plan_rsya) * 100 if lead_price_plan_rsya else 0

                st.markdown(f"""
                **РСЯ:**
                - Выполнение плана по расходу: {rsya_cost / rsya_cost_plan * 100:.1f}%
                - Выполнение плана по показам: {rsya_impressions / rsya_impressions_plan * 100:.1f}%
                - Выполнение плана по кликам: {rsya_clicks / rsya_clicks_plan * 100:.1f}%
                - Выполнение плана по заявкам: {rsya_conversions / rsya_conversions_plan * 100:.1f}%
                - Цена лида (факт): {lead_price_fact_rsya:.0f} ₽
                - Цена лида (план): {lead_price_plan_rsya:.0f} ₽
                - Отклонение цены лида от плана: {lead_price_diff_pct_rsya:+.1f}%
                """)

        except Exception as e:
            st.error(f"❌ Ошибка обработки: {e}")

if not result_df.empty:
    st.success(f"✅ Найдено совпадений: {len(result_df)}")
